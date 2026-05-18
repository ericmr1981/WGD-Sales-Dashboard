"""File import utilities: parse XLSX/CSV and upload to Supabase."""
import csv
import json
import ssl
import io
import urllib.parse
import urllib.request
from typing import Any, List, Optional

import openpyxl

from db import _get_supabase_config

_ctx = ssl.create_default_context()
_ctx.check_hostname = False
_ctx.verify_mode = ssl.CERT_NONE


def clean_value(val: Any) -> Optional[str]:
    if val is None:
        return None
    val = str(val).strip().strip("`").strip("'")
    return val if val else None


def _supabase_request(method: str, path: str, data: Optional[list] = None):
    supabase_url, supabase_key = _get_supabase_config()
    if not supabase_url or not supabase_key:
        raise RuntimeError("SUPABASE_URL and SUPABASE_KEY must be configured.")
    url = f"{supabase_url}/rest/v1/{path}"
    body = json.dumps(data).encode() if data else None
    req = urllib.request.Request(url, data=body, method=method)
    req.add_header("apikey", supabase_key)
    req.add_header("Authorization", f"Bearer {supabase_key}")
    req.add_header("Content-Type", "application/json")
    req.add_header("Accept", "application/json")
    if method == "POST":
        req.add_header("Prefer", "resolution=merge-duplicates")
    try:
        resp = urllib.request.urlopen(req, context=_ctx)
        raw = resp.read().decode()
        return json.loads(raw) if raw else []
    except urllib.error.HTTPError as e:
        body = e.read().decode()
        if e.code == 409:
            return None
        raise RuntimeError(f"Supabase HTTP {e.code}: {body[:200]}")
    except Exception as e:
        raise RuntimeError(f"Supabase request failed: {e}")


CHANNEL_HEADER_MAP = {
    "免支付": "free_payment",
    "微信支付": "wechat_pay",
    "支付宝支付": "alipay",
    "现金支付": "cash",
    "美团团购券": "meituan_coupon",
    "抖音团购券": "douyin_coupon",
    "云闪付": "yunshanfu",
    "自定义结账方式": "custom_payment",
}


def parse_pos_orders(file_bytes: bytes) -> List[dict]:
    """Parse 收银明细表 XLSX into list of row dicts for pos_orders table.

    Reads row 2 sub-headers to determine which payment channel is in which column,
    since column positions vary across different XLSX files.
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active

    # Build channel column map from row 2 sub-headers
    channel_cols = {}
    for col_idx in range(1, ws.max_column + 1):
        header = clean_value(ws.cell(2, col_idx).value)
        if header and header in CHANNEL_HEADER_MAP:
            channel_cols[CHANNEL_HEADER_MAP[header]] = col_idx

    def _chan(db_col: str, row_idx: int) -> float:
        c = channel_cols.get(db_col)
        if c is None:
            return 0.0
        v = ws.cell(row_idx, c).value
        return float(v) if v is not None else 0.0

    rows = []
    seen = set()
    for row_idx in range(3, ws.max_row + 1):
        order_no = clean_value(ws.cell(row_idx, 3).value)
        if not order_no or order_no in seen:
            continue
        seen.add(order_no)
        rows.append({
            "order_no": order_no,
            "store_name": clean_value(ws.cell(row_idx, 1).value),
            "sale_date": clean_value(ws.cell(row_idx, 2).value),
            "total_revenue": _chan("_total", row_idx) if "_total" in channel_cols
                else float(ws.cell(row_idx, 4).value or 0),
            "gross_income": float(ws.cell(row_idx, 5).value or 0),
            "discount_total": float(ws.cell(row_idx, 6).value or 0),
            "net_revenue": float(ws.cell(row_idx, 7).value or 0),
            "quantity": int(float(ws.cell(row_idx, 8).value or 0)),
            "free_payment": _chan("free_payment", row_idx),
            "wechat_pay": _chan("wechat_pay", row_idx),
            "douyin_coupon": _chan("douyin_coupon", row_idx),
            "alipay": _chan("alipay", row_idx),
            "cash": _chan("cash", row_idx),
            "meituan_coupon": _chan("meituan_coupon", row_idx),
            "custom_payment": _chan("custom_payment", row_idx),
            "yunshanfu": _chan("yunshanfu", row_idx),
        })
    return rows


def parse_revenue_csv(file_bytes: bytes) -> List[dict]:
    """Parse 收入明细表 CSV into list of row dicts for revenue_detail table."""
    text = file_bytes.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    rows = []
    for row in reader:
        order_no = clean_value(row.get("订单号", ""))
        if not order_no:
            continue
        rows.append({
            "order_no": order_no,
            "brand_name": clean_value(row.get("品牌名称", "")),
            "city": clean_value(row.get("城市", "")),
            "store_name": clean_value(row.get("门店名称", "")),
            "store_id": clean_value(row.get("门店id", "")),
            "store_code": clean_value(row.get("门店编码", "")),
            "sale_date": clean_value(row.get("营业日期", "")),
            "order_time": clean_value(row.get("下单时间", "")),
            "payment_time": clean_value(row.get("支付时间", "")),
            "third_party_order_no": clean_value(row.get("三方订单号", "")),
            "third_party_payment_no": clean_value(row.get("三方支付流水号", "")),
            "merchant_order_no": clean_value(row.get("商户订单号", "")),
            "third_party_coupon_id": clean_value(row.get("三方券id", "")),
            "order_source": clean_value(row.get("订单来源", "")),
            "order_type": clean_value(row.get("订单类型", "")),
            "order_status": clean_value(row.get("订单状态", "")),
            "is_reversed": clean_value(row.get("是否反结", "否")) or "否",
            "member_id": clean_value(row.get("会员id", "")),
            "user_phone": clean_value(row.get("用户手机号", "")),
            "payment_split": clean_value(row.get("结账方式拆分", "")),
            "payment_method": clean_value(row.get("结账方式名称", "")),
            "store_open_date": clean_value(row.get("门店开业日期", "")),
            "gross_income": float(clean_value(row.get("营业收入", "0")) or 0),
            "net_revenue": float(clean_value(row.get("营业净收", "0")) or 0),
            "overflow_amount": float(clean_value(row.get("溢收金额", "0")) or 0),
            "total_revenue": float(clean_value(row.get("营业额", "0")) or 0),
            "discount_total": float(clean_value(row.get("优惠总额", "0")) or 0),
            "coupon_service_fee": float(clean_value(row.get("团购券手续费", "0")) or 0),
        })
    return rows


def parse_product_sales(file_bytes: bytes) -> List[dict]:
    """Parse 商品明细表 CSV into list of row dicts for product_sales table."""
    text = file_bytes.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    rows = []
    for row in reader:
        name = clean_value(row.get("商品名称", ""))
        if not name or name == "--":
            continue
        rows.append({
            "store_name": clean_value(row.get("门店名称", "")),
            "sale_date": clean_value(row.get("日期", "")),
            "order_no": clean_value(row.get("订单号", "")),
            "product_name": name,
            "unit_price": float(clean_value(row.get("商品原价", "0")) or 0),
            "quantity": int(float(clean_value(row.get("销售数量", "1")) or 1)),
            "total_price": float(clean_value(row.get("商品销售额", "0")) or 0),
            "actual_received": float(clean_value(row.get("商品实收", "0")) or 0),
            "discount": float(clean_value(row.get("商品优惠", "0")) or 0),
        })
    return rows


DEDUP_BATCH = 500


def check_existing_orders(order_nos: List[str]) -> set:
    """Query Supabase for which order_nos already exist in pos_orders."""
    if not order_nos:
        return set()
    existing = set()
    for i in range(0, len(order_nos), DEDUP_BATCH):
        chunk = order_nos[i:i + DEDUP_BATCH]
        quoted = ",".join(urllib.parse.quote(no, safe="") for no in chunk)
        result = _supabase_request("GET", f"pos_orders?select=order_no&order_no=in.({quoted})")
        existing.update(r["order_no"] for r in result if r.get("order_no"))
    return existing


def check_existing_revenue(order_nos: List[str]) -> set:
    """Query Supabase for which order_nos already exist in revenue_detail."""
    if not order_nos:
        return set()
    existing = set()
    for i in range(0, len(order_nos), DEDUP_BATCH):
        chunk = order_nos[i:i + DEDUP_BATCH]
        quoted = ",".join(urllib.parse.quote(no, safe="") for no in chunk)
        result = _supabase_request("GET", f"revenue_detail?select=order_no&order_no=in.({quoted})")
        existing.update(r["order_no"] for r in result if r.get("order_no"))
    return existing


def check_existing_product_sales(keys: List[tuple]) -> set:
    """Query Supabase for which (order_no, product_name) pairs already exist.

    Keys is a list of (order_no, product_name) tuples.
    Batches queries to stay within URL length limits.
    """
    if not keys:
        return set()
    order_nos = list({k[0] for k in keys})
    existing = set()
    for i in range(0, len(order_nos), DEDUP_BATCH):
        chunk = order_nos[i:i + DEDUP_BATCH]
        quoted = ",".join(urllib.parse.quote(no, safe="") for no in chunk)
        result = _supabase_request(
            "GET",
            f"product_sales?select=order_no,product_name&order_no=in.({quoted})"
        )
        existing.update(
            (r["order_no"], r["product_name"])
            for r in result
            if r.get("order_no") and r.get("product_name")
        )
    return existing


def upload_batch(table: str, rows: List[dict]) -> tuple:
    """Upload rows to Supabase table in batches of 500. Returns (uploaded, total)."""
    total = 0
    for i in range(0, len(rows), 500):
        chunk = rows[i:i + 500]
        _supabase_request("POST", table, chunk)
        total += len(chunk)
    return total, len(rows)
