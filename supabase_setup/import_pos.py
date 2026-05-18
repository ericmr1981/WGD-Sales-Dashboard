"""Import 收银明细表 (xlsx) and 商品销售明细表 (csv) to Supabase."""
import csv
import json
import os
import ssl
import urllib.request
import urllib.parse

SUPABASE_URL = os.environ.get("SUPABASE_URL", "")
SUPABASE_KEY = os.environ.get("SUPABASE_KEY", "")
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

_ctx = ssl.create_default_context()
_ctx.check_hostname = False
_ctx.verify_mode = ssl.CERT_NONE


def supabase_request(method, path, data=None):
    url = f"{SUPABASE_URL}/rest/v1/{path}"
    body = json.dumps(data).encode() if data else None
    req = urllib.request.Request(url, data=body, method=method)
    req.add_header("apikey", SUPABASE_KEY)
    req.add_header("Authorization", f"Bearer {SUPABASE_KEY}")
    req.add_header("Content-Type", "application/json")
    req.add_header("Accept", "application/json")
    if method == "POST":
        req.add_header("Prefer", "resolution=merge-duplicates")
    try:
        resp = urllib.request.urlopen(req, context=_ctx)
        raw = resp.read().decode()
        if raw:
            return json.loads(raw)
        return []
    except urllib.error.HTTPError as e:
        body = e.read().decode()
        if e.code == 409:
            return None
        print(f"  HTTP {e.code}: {body[:200]}")
        return None
    except Exception as e:
        print(f"  请求失败: {e}")
        return None


def clean_value(val):
    if val is None:
        return None
    val = str(val).strip().strip("`").strip("'")
    return val if val else None


CHANNEL_HEADER_MAP = {
    "免支付": "free_payment",
    "微信支付": "wechat_pay",
    "支付宝支付": "alipay",
    "现金支付": "cash",
    "美团团购券": "meituan_coupon",
    "抖音团购券": "douyin_coupon",
    "云闪付": "yunshanfu",
    "自定义结账方式": "custom_payment",
}


def import_pos_orders():
    """Import 收银明细表 xlsx — reads row 2 headers to map payment channels by name."""
    import openpyxl

    filepath = os.path.join(BASE_DIR, "收银明细表-2026-04-012026-04-30-9d15c166a917437b9cbb8f575ba0ce2c.xlsx")
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active

    # Map channel header name → column index from row 2
    chan_cols = {}
    for c in range(1, ws.max_column + 1):
        h = clean_value(ws.cell(2, c).value)
        if h and h in CHANNEL_HEADER_MAP:
            chan_cols[CHANNEL_HEADER_MAP[h]] = c

    def chan_val(db_col, row_idx):
        c = chan_cols.get(db_col)
        if c is None:
            return 0.0
        v = ws.cell(row_idx, c).value
        return float(v) if v is not None else 0.0

    batch = []
    seen = set()
    for row_idx in range(3, ws.max_row + 1):
        order_no = clean_value(ws.cell(row_idx, 3).value)
        if not order_no:
            continue
        if order_no in seen:
            continue
        seen.add(order_no)

        batch.append({
            "order_no": order_no,
            "store_name": clean_value(ws.cell(row_idx, 1).value),
            "sale_date": clean_value(ws.cell(row_idx, 2).value),
            "total_revenue": float(ws.cell(row_idx, 4).value or 0),
            "gross_income": float(ws.cell(row_idx, 5).value or 0),
            "discount_total": float(ws.cell(row_idx, 6).value or 0),
            "net_revenue": float(ws.cell(row_idx, 7).value or 0),
            "quantity": int(float(ws.cell(row_idx, 8).value or 0)),
            "free_payment": chan_val("free_payment", row_idx),
            "wechat_pay": chan_val("wechat_pay", row_idx),
            "douyin_coupon": chan_val("douyin_coupon", row_idx),
            "alipay": chan_val("alipay", row_idx),
            "cash": chan_val("cash", row_idx),
            "meituan_coupon": chan_val("meituan_coupon", row_idx),
            "custom_payment": chan_val("custom_payment", row_idx),
            "yunshanfu": chan_val("yunshanfu", row_idx),
        })

    print(f"pos_orders: 共 {len(batch)} 条（去重后）")
    for i in range(0, len(batch), 500):
        chunk = batch[i:i + 500]
        supabase_request("POST", "pos_orders", chunk)
        print(f"  pos_orders: 已导入 {min(i + 500, len(batch))}/{len(batch)}")
    print(f"pos_orders: 完成，共 {len(batch)} 条")


def import_product_sales():
    filepath = os.path.join(BASE_DIR, "商品销售明细表 2026年4月 (1).csv")
    with open(filepath, mode="r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        batch = []
        for row in reader:
            name = clean_value(row.get("商品名称", ""))
            if name == "--" or not name:
                continue
            batch.append({
                "store_name": clean_value(row.get("门店名称", "")),
                "sale_date": clean_value(row.get("日期", "")),
                "order_no": clean_value(row.get("订单号", "")),
                "product_name": name,
                "unit_price": float(clean_value(row.get("商品原价", "0")) or 0),
                "quantity": int(float(clean_value(row.get("销售数量", "1")) or 1)),
                "total_price": float(clean_value(row.get("商品销售额", "0")) or 0),
                "actual_received": float(clean_value(row.get("商品实收", "0")) or 0),
                "discount": float(clean_value(row.get("商品优惠", "0")) or 0),
            })

    print(f"product_sales: 共 {len(batch)} 条")
    for i in range(0, len(batch), 500):
        chunk = batch[i:i + 500]
        supabase_request("POST", "product_sales", chunk)
        print(f"  product_sales: 已导入 {min(i + 500, len(batch))}/{len(batch)}")
    print(f"product_sales: 完成，共 {len(batch)} 条")


if __name__ == "__main__":
    print("开始导入 pos_orders...")
    import_pos_orders()
    print()
    print("开始导入 product_sales...")
    import_product_sales()
    print()
    print("全部导入完成！")